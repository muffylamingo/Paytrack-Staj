"""
routers/invoices.py — Fatura endpoint'leri (HTTP katmanı).

PDF 2. + 4. + 5. Aşama işlemleri:
  POST   /invoices           -> yeni fatura
  GET    /invoices           -> listele (+ filtreler + sıralama)
  GET    /invoices/export    -> Excel (.xlsx) indir
  GET    /invoices/{id}      -> tek fatura
  PUT    /invoices/{id}      -> güncelle
  PATCH  /invoices/{id}/pay  -> "Ödendi" işaretle
  DELETE /invoices/{id}      -> sil

Ekstra (Özellik #4 — dosya eki):
  POST   /invoices/{id}/attachment -> dekont/PDF yükle
  GET    /invoices/{id}/attachment -> ekli dosyayı indir
  DELETE /invoices/{id}/attachment -> eki kaldır

Ekstra (Özellik #5 — tekrarlayan faturalar):
  POST   /invoices/generate-recurring -> eksik tekrarları üret

Ekstra (Özellik #7 — Excel'den içe aktarma):
  GET    /invoices/import-template -> boş şablonu indir
  POST   /invoices/import          -> doldurulmuş dosyadan toplu yükle
"""
import io
from typing import Literal

from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pydantic import ValidationError
from sqlalchemy.orm import Session

from app import crud, excel_io, models, schemas, storage
from app.database import get_db

router = APIRouter(prefix="/invoices", tags=["Faturalar"])


@router.post("", response_model=schemas.InvoiceOut, status_code=201)
def create_invoice(data: schemas.InvoiceCreate, db: Session = Depends(get_db)):
    """Yeni fatura ekler. (201 = Created)"""
    return crud.create_invoice(db, data)


@router.get("", response_model=list[schemas.InvoiceOut])
def list_invoices(
    status: schemas.InvoiceStatus | None = Query(default=None, description="Duruma göre süz"),
    category: schemas.Category | None = Query(default=None, description="Kategoriye göre süz"),
    vendor: str | None = Query(default=None, description="Tedarikçi adında ara"),
    sort: Literal[
        "due_date", "amount", "vendor_name", "invoice_number", "status", "category"
    ] = Query(default="due_date", description="Sıralama alanı"),
    order: Literal["asc", "desc"] = Query(default="asc", description="Sıralama yönü"),
    db: Session = Depends(get_db),
):
    """Faturaları listeler. Filtreler: ?status= ?category= ?vendor= · Sıralama: ?sort= ?order="""
    invoices = crud.get_invoices(db, status=status, category=category, vendor=vendor, sort=sort, order=order)
    # Her faturaya güncel kurla TL karşılığını ekle (Ekstra #8)
    return crud.attach_try_amounts(db, invoices)


# NOT: Bu route, aşağıdaki /{invoice_id} route'undan ÖNCE tanımlanmalı
# ki "export" bir id sanılmasın.
@router.get("/export")
def export_invoices(
    status: schemas.InvoiceStatus | None = Query(default=None),
    category: schemas.Category | None = Query(default=None),
    vendor: str | None = Query(default=None),
    sort: Literal[
        "due_date", "amount", "vendor_name", "invoice_number", "status", "category"
    ] = Query(default="due_date"),
    order: Literal["asc", "desc"] = Query(default="asc"),
    db: Session = Depends(get_db),
):
    """Faturaları (mevcut filtre/sıralama ile) Excel (.xlsx) dosyası olarak indirir."""
    invoices = crud.get_invoices(db, status=status, category=category, vendor=vendor, sort=sort, order=order)

    wb = Workbook()
    ws = wb.active
    ws.title = "Faturalar"

    headers = ["Fatura No", "Tedarikçi", "Kategori", "Tutar", "Döviz", "TL Karşılığı",
               "Son Ödeme", "Durum", "Not"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)  # başlık satırını kalın yap

    rates = crud.get_rates_map(db)   # TL karşılığı için güncel kurlar
    for inv in invoices:
        ws.append([
            inv.invoice_number,
            inv.vendor_name,
            inv.category,
            float(inv.amount),
            inv.currency,
            float(crud.to_try(inv.amount, inv.currency, rates)),
            inv.due_date.isoformat(),
            inv.status,
            inv.notes or "",
        ])

    # Kolon genişlikleri (okunabilirlik)
    for i, width in enumerate([16, 22, 12, 12, 8, 14, 14, 12, 30], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="faturalar.xlsx"'},
    )


@router.get("/import-template")
def import_template():
    """Boş içe aktarma şablonunu (.xlsx) indirir — kullanıcı doğru sütunları görsün."""
    return Response(
        content=excel_io.build_template(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="fatura-sablonu.xlsx"'},
    )


@router.post("/import", response_model=schemas.ImportResult)
async def import_invoices(
    file: UploadFile = File(..., description="Doldurulmuş .xlsx dosyası"),
    db: Session = Depends(get_db),
):
    """
    Excel dosyasından toplu fatura yükler.

    KISMİ BAŞARI stratejisi: geçerli satırlar eklenir, hatalı satırlar atlanır ve
    kullanıcıya satır numarasıyla raporlanır. (Alternatif "hepsi ya da hiçbiri"
    olurdu; 500 satırın 1'i bozuk diye 499'unu geri çevirmek kullanıcıyı yorar.)
    """
    if file.content_type not in excel_io.ALLOWED_TYPES and not (file.filename or "").endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Sadece .xlsx dosyası yüklenebilir.")

    content = await file.read()
    try:
        rows = excel_io.parse_invoices(content)
    except excel_io.ExcelError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e

    mevcut = crud.existing_invoice_numbers(db)   # veritabanındaki numaralar
    dosyada: set[str] = set()                    # dosyanın kendi içindeki kopyalar
    hazir: list[models.Invoice] = []
    hatalar: list[dict] = []

    for row_no, data in rows:
        # 1) Hücre çevirme hatası (tarih/tutar okunamadı)
        if "__error__" in data:
            hatalar.append({"row": row_no, "message": data["__error__"]})
            continue

        # 2) Kopya kontrolü — aynı fatura numarası iki kez girilmesin
        numara = data.get("invoice_number")
        if numara in mevcut:
            hatalar.append({"row": row_no, "message": f"'{numara}' zaten kayıtlı, atlandı"})
            continue
        if numara in dosyada:
            hatalar.append({"row": row_no, "message": f"'{numara}' dosyada tekrar ediyor, atlandı"})
            continue

        # 3) Asıl doğrulama: Pydantic şemasından geçir (kategori/durum/tutar>0 ...)
        try:
            gecerli = schemas.InvoiceCreate(**data)
        except ValidationError as e:
            ilk = e.errors()[0]
            alan = ilk["loc"][0] if ilk["loc"] else "?"
            hatalar.append({"row": row_no, "message": f"{alan}: {ilk['msg']}"})
            continue

        dosyada.add(numara)
        hazir.append(models.Invoice(**gecerli.model_dump()))

    crud.bulk_create_invoices(db, hazir)

    return {
        "imported": len(hazir),
        "failed": len(hatalar),
        "errors": hatalar[:20],   # raporu şişirmemek için ilk 20 hata
    }


# NOT: Bu da /{invoice_id}'den ÖNCE — yol adları id sanılmasın.
@router.post("/generate-recurring", response_model=schemas.GenerateResult)
def generate_recurring(db: Session = Depends(get_db)):
    """
    Tekrarlayan faturaların eksik tekrarlarını üretir (bugünden 30 gün ilerisine kadar).
    Tekrar tekrar çağrılabilir: zaten üretilmişse yenisini oluşturmaz (idempotent).
    """
    created = crud.generate_recurring(db)
    return {"created": len(created), "invoices": created}


@router.get("/{invoice_id}", response_model=schemas.InvoiceOut)
def get_invoice(invoice_id: int, db: Session = Depends(get_db)):
    """Tek bir faturayı getirir."""
    invoice = crud.get_invoice(db, invoice_id)
    if not invoice:
        raise HTTPException(status_code=404, detail="Fatura bulunamadı")
    return invoice


@router.put("/{invoice_id}", response_model=schemas.InvoiceOut)
def update_invoice(invoice_id: int, data: schemas.InvoiceUpdate, db: Session = Depends(get_db)):
    """Faturayı günceller (sadece gönderilen alanlar)."""
    invoice = crud.get_invoice(db, invoice_id)
    if not invoice:
        raise HTTPException(status_code=404, detail="Fatura bulunamadı")
    return crud.update_invoice(db, invoice, data)


@router.patch("/{invoice_id}/pay", response_model=schemas.InvoiceOut)
def pay_invoice(invoice_id: int, db: Session = Depends(get_db)):
    """Faturayı 'Ödendi' olarak işaretler."""
    invoice = crud.get_invoice(db, invoice_id)
    if not invoice:
        raise HTTPException(status_code=404, detail="Fatura bulunamadı")
    return crud.mark_paid(db, invoice)


@router.delete("/{invoice_id}", status_code=204)
def delete_invoice(invoice_id: int, db: Session = Depends(get_db)):
    """Faturayı siler. (204 = No Content)"""
    invoice = crud.get_invoice(db, invoice_id)
    if not invoice:
        raise HTTPException(status_code=404, detail="Fatura bulunamadı")
    crud.delete_invoice(db, invoice)


# ---------------------------------------------------------------
# Ek dosya (dekont/fatura PDF'i veya görseli)  — Ekstra Özellik #4
# ---------------------------------------------------------------
def _get_or_404(db: Session, invoice_id: int) -> models.Invoice:
    """Tekrar eden 404 kontrolünü tek yerde topladık."""
    invoice = crud.get_invoice(db, invoice_id)
    if not invoice:
        raise HTTPException(status_code=404, detail="Fatura bulunamadı")
    return invoice


@router.post("/{invoice_id}/attachment", response_model=schemas.InvoiceOut)
async def upload_attachment(
    invoice_id: int,
    file: UploadFile = File(..., description="PDF/JPG/PNG/WEBP — en fazla 5 MB"),
    db: Session = Depends(get_db),
):
    """Faturaya dekont/fatura dosyası yükler. Zaten ek varsa üzerine yazar."""
    invoice = _get_or_404(db, invoice_id)
    content = await file.read()   # dosya içeriğini belleğe al (5 MB sınırı var)
    try:
        stored_name = storage.save_upload(content, file.content_type)
    except storage.UploadError as e:
        # 400 = Bad Request (kullanıcı hatası); 500 değil, çünkü sunucu bozuk değil
        raise HTTPException(status_code=400, detail=str(e)) from e
    return crud.set_attachment(db, invoice, stored_name, file.filename or "dosya")


@router.get("/{invoice_id}/attachment")
def download_attachment(invoice_id: int, db: Session = Depends(get_db)):
    """Faturaya ekli dosyayı indirir (orijinal adıyla)."""
    invoice = _get_or_404(db, invoice_id)
    if not invoice.attachment_path:
        raise HTTPException(status_code=404, detail="Bu faturada ek dosya yok")

    path = storage.file_path(invoice.attachment_path)
    if not path.exists():
        # Veritabanında kayıt var ama dosya diskte yok (elle silinmiş olabilir)
        raise HTTPException(status_code=404, detail="Dosya diskte bulunamadı")

    return FileResponse(path, filename=invoice.attachment_name or path.name)


@router.delete("/{invoice_id}/attachment", response_model=schemas.InvoiceOut)
def remove_attachment(invoice_id: int, db: Session = Depends(get_db)):
    """Faturanın ek dosyasını kaldırır (fatura kalır, sadece dosya silinir)."""
    invoice = _get_or_404(db, invoice_id)
    if not invoice.attachment_path:
        raise HTTPException(status_code=404, detail="Bu faturada ek dosya yok")
    return crud.clear_attachment(db, invoice)
