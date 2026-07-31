"""
excel_io.py — Excel dosyasını OKUMA (içe aktarma) ve şablon üretme.

Dışa aktarma (export) hâlâ routers/invoices.py içinde; burası okuma tarafı.
storage.py gibi bu da tek bir işten sorumlu: "Excel ↔ Python sözlüğü" çevirisi.
Veritabanına hiç dokunmaz.

İÇE AKTARMA, DIŞA AKTARMADAN ÇOK DAHA ZORDUR:
dışa aktarırken veriye biz hükmederiz; içe aktarırken kullanıcının dosyası gelir —
sütunlar eksik/karışık olabilir, tarih "15.08.2026" yazılmış olabilir, tutar
"4.200,50" gibi Türkçe biçimde gelebilir, satır bomboş olabilir...
"""
import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook

MAX_ROWS = 1000  # emniyet freni: devasa dosya sunucuyu kilitlemesin

ALLOWED_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
    "application/vnd.ms-excel",  # bazı tarayıcılar .xlsx'i böyle etiketler
}

# Excel başlığı -> bizim alan adımız. Türkçe karakterli ve karaktersiz
# yazımların ikisini de kabul ediyoruz (kullanıcıya kolaylık).
HEADER_MAP = {
    "fatura no": "invoice_number",
    "fatura numarası": "invoice_number",
    "tedarikçi": "vendor_name",
    "tedarikci": "vendor_name",
    "kategori": "category",
    "tutar": "amount",
    "döviz": "currency",
    "doviz": "currency",
    "para birimi": "currency",
    "son ödeme": "due_date",
    "son odeme": "due_date",
    "son ödeme tarihi": "due_date",
    "durum": "status",
    "not": "notes",
    "açıklama": "notes",
}

REQUIRED = {"invoice_number", "vendor_name", "category", "amount", "due_date"}

TEMPLATE_HEADERS = ["Fatura No", "Tedarikçi", "Kategori", "Tutar", "Döviz", "Son Ödeme", "Durum", "Not"]


class ExcelError(Exception):
    """Dosyanın tamamıyla ilgili hata (bozuk dosya, başlık yok...) -> 400."""


def _parse_amount(value) -> Decimal:
    """Tutarı Decimal'e çevirir. '4.200,50' (TR) ve '4200.50' (EN) biçimlerini de anlar."""
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).strip().replace(" ", "").replace("₺", "")
    # Türkçe biçim: nokta binlik ayıracı, virgül ondalık -> "4.200,50" => "4200.50"
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation as e:
        raise ValueError(f"Tutar sayıya çevrilemedi: '{value}'") from e


def _parse_date(value) -> date:
    """Tarihi date'e çevirir. Excel tarih hücresi, ISO metin ve 15.08.2026 kabul edilir."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Tarih anlaşılamadı: '{value}' (örn. 2026-08-15 veya 15.08.2026)")


def build_template() -> bytes:
    """Kullanıcının dolduracağı boş şablonu (örnek satırla birlikte) üretir."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Faturalar"
    ws.append(TEMPLATE_HEADERS)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    # Örnek satır — kullanıcı biçimi görsün diye
    ws.append(["FTR-2026-200", "Örnek Tedarikçi", "Enerji", 1500.50, "TRY", "2026-09-15", "Bekliyor", "örnek not"])
    for i, width in enumerate([16, 22, 12, 12, 8, 14, 12, 30], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def parse_invoices(content: bytes) -> list[tuple[int, dict]]:
    """
    Excel içeriğini okuyup (satır_numarası, alan sözlüğü) listesi döndürür.

    Doğrulama YAPMAZ (o Pydantic'in işi) — sadece hücreleri okur ve tipleri çevirir.
    Çeviremediği hücre için sözlüğe hata mesajı koyar ki router satırı raporlayabilsin.
    """
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as e:  # bozuk/xlsx olmayan dosya
        raise ExcelError("Dosya okunamadı. Geçerli bir .xlsx dosyası mı?") from e

    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration:
        raise ExcelError("Dosya boş.") from None

    # Başlıkları eşle: {sütun indeksi: alan adı}
    columns: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        key = str(cell or "").strip().lower()
        if key in HEADER_MAP:
            columns[idx] = HEADER_MAP[key]

    missing = REQUIRED - set(columns.values())
    if missing:
        adlar = {
            "invoice_number": "Fatura No", "vendor_name": "Tedarikçi", "category": "Kategori",
            "amount": "Tutar", "due_date": "Son Ödeme",
        }
        raise ExcelError(
            "Şu sütunlar bulunamadı: " + ", ".join(sorted(adlar[m] for m in missing))
            + ". Şablonu indirip kullanabilirsin."
        )

    result: list[tuple[int, dict]] = []
    for row_no, row in enumerate(rows, start=2):  # 2 = başlıktan sonraki ilk satır
        if len(result) >= MAX_ROWS:
            raise ExcelError(f"Çok fazla satır (en fazla {MAX_ROWS}).")
        # Tamamen boş satırları sessizce atla (Excel'de sık olur)
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        data: dict = {}
        for idx, field in columns.items():
            value = row[idx] if idx < len(row) else None
            if value is None or str(value).strip() == "":
                continue
            try:
                if field == "amount":
                    data[field] = _parse_amount(value)
                elif field == "due_date":
                    data[field] = _parse_date(value)
                else:
                    data[field] = str(value).strip()
            except ValueError as e:
                data["__error__"] = str(e)

        result.append((row_no, data))

    return result
